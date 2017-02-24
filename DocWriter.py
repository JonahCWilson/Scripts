from openpyxl import *

data = open('studentdata.txt', 'r')
#base = load_workbook('student.xlsx')
header = '---'
lines = map(lambda a: a.strip('\n'), data.readlines())
students = []
curr = []
letters = 'ABCD'

for line in lines:
    if line == header:
	students.append(curr)
	curr = []
    else:
	curr.append(line)

for s in students:
    start = 23
    last, first, id, school, grade = s[0].split(',')
    nameString = last + ', ' + first + ' ' + id
    wb = load_workbook('Student.xlsx')
    ws = wb.active
    ws['B4'] = nameString
    ws['B5'] = school
    ws['B6'] = grade
    for line in s[1:]:
	info = line.split('/')
	for i in range(4):
	    ws[letters[i] + str(start)] = info[i]
	start += 1
    wb.save(nameString + '.xlsx')
